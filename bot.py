import os
import asyncio
import httpx
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, types, F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, FSInputFile, ReplyKeyboardMarkup, KeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder

# Загрузка переменных окружения
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = 6201234513
FIREBASE_URL = "https://neonapp-a05b0-default-rtdb.firebaseio.com/"

# Инициализация бота
bot = Bot(token=BOT_TOKEN)
router = Router()

# Состояния FSM
class Form(StatesGroup):
    enter_pin = State()
    add_supplier_name = State()
    edit_supplier_name = State()
    add_debt_supplier = State()
    add_debt_amount = State()
    add_debt_date = State()
    add_payment_supplier = State()
    add_payment_amount = State()
    add_payment_date = State()

# --- РАБОТА С БАЗОЙ ДАННЫХ (Direct HTTP) ---
async def fb_get(path=""):
    async with httpx.AsyncClient() as client:
        r = await client.get(f"{FIREBASE_URL}{path}.json")
        return r.json() or {}

async def fb_put(path, data):
    async with httpx.AsyncClient() as client:
        await client.put(f"{FIREBASE_URL}{path}.json", json=data)

async def fb_post(path, data):
    async with httpx.AsyncClient() as client:
        r = await client.post(f"{FIREBASE_URL}{path}.json", json=data)
        return r.json()

async def fb_delete(path):
    async with httpx.AsyncClient() as client:
        await client.delete(f"{FIREBASE_URL}{path}.json")

async def is_user_allowed(user_id):
    users = await fb_get(f"allowed_users/{user_id}")
    return users == True or str(user_id) == str(ADMIN_ID)

# Клавиатуры
def main_reply_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.row(KeyboardButton(text="📦 Поставщики"), KeyboardButton(text="➕ Новый поставщик"))
    builder.row(KeyboardButton(text="💸 Активные долги"), KeyboardButton(text="➕ Новый долг"))
    builder.row(KeyboardButton(text="💳 Внести оплату"), KeyboardButton(text="📊 Отчет Excel"))
    builder.row(KeyboardButton(text="⚙️ Управление"))
    return builder.as_markup(resize_keyboard=True)

def main_menu():
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📦 Поставщики", callback_data="view_suppliers"),
                InlineKeyboardButton(text="➕ Новый поставщик", callback_data="add_supplier"))
    builder.row(InlineKeyboardButton(text="💸 Долги", callback_data="view_debts"),
                InlineKeyboardButton(text="➕ Новый долг", callback_data="add_debt"))
    builder.row(InlineKeyboardButton(text="📊 Экспорт в Excel", callback_data="export_excel"))
    builder.row(InlineKeyboardButton(text="⚠️ Очистить базу", callback_data="confirm_clear"))
    return builder.as_markup()

# Обработчики команд
@router.message(Command("start"))
async def start_cmd(message: types.Message):
    user_id = message.from_user.id
    
    # Если это админ или разрешенный пользователь
    if await is_user_allowed(user_id):
        await message.answer("👋 С возвращением! Выберите действие:", 
                             reply_markup=main_reply_keyboard())
        return

    # Если пользователь новый - отправляем запрос админу
    await message.answer("⏳ Запрос на доступ отправлен администратору. Ожидайте подтверждения...")
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✅ Принять", callback_data=f"auth_accept_{user_id}"),
        InlineKeyboardButton(text="❌ Отклонить", callback_data=f"auth_reject_{user_id}")
    )
    
    user_info = f"👤 **Новый запрос доступа:**\nID: `{user_id}`\nИмя: {message.from_user.full_name}\nUsername: @{message.from_user.username}"
    await bot.send_message(chat_id=ADMIN_ID, text=user_info, reply_markup=builder.as_markup(), parse_mode="Markdown")

# Обработка авторизации админом
@router.callback_query(F.data.startswith("auth_"))
async def process_auth(callback: types.CallbackQuery):
    if str(callback.from_user.id) != str(ADMIN_ID):
        await callback.answer("Только администратор может это делать!", show_alert=True)
        return

    parts = callback.data.split("_")
    action = parts[1]  # 'accept' or 'reject'
    user_id = parts[2]
    
    if action == "accept":
        await fb_put(f"allowed_users/{user_id}", True)
        await bot.send_message(chat_id=user_id, text="✅ Доступ разрешен! Используйте кнопки меню:", reply_markup=main_reply_keyboard())
        await callback.message.edit_text(f"✅ Пользователь {user_id} допущен.")
    else:
        await bot.send_message(chat_id=user_id, text="❌ К сожалению, вам отказано в доступе.")
        await callback.message.edit_text(f"❌ Пользователю {user_id} отказано.")
    await callback.answer()

# Обработка Reply-кнопок
@router.message(F.text == "📦 Поставщики")
async def btn_suppliers(message: types.Message):
    await view_suppliers_msg(message)

@router.message(F.text == "💸 Активные долги")
async def btn_debts_alias(message: types.Message):
    await view_debts_msg(message)

@router.message(F.text.contains("Активные долги"))
async def btn_debts_contains(message: types.Message):
    await view_debts_msg(message)

@router.message(F.text == "📊 Отчет Excel")
async def btn_report(message: types.Message):
    await export_to_excel_msg(message)

@router.message(F.text == "➕ Новый долг")
async def btn_add_debt(message: types.Message, state: FSMContext):
    await add_debt_start_msg(message, state)

@router.message(F.text == "➕ Новый поставщик")
async def btn_add_supplier(message: types.Message, state: FSMContext):
    await add_supplier_start_msg(message, state)

@router.message(F.text == "💳 Внести оплату")
async def btn_add_payment(message: types.Message, state: FSMContext):
    await add_payment_start_msg(message, state)

@router.message(F.text == "⚙️ Управление")
async def btn_management_start(message: types.Message, state: FSMContext):
    await message.answer("🔒 Введите ПИН-код для доступа к управлению:")
    await state.set_state(Form.enter_pin)

@router.message(Form.enter_pin)
async def process_pin(message: types.Message, state: FSMContext):
    if message.text == "88":
        await state.clear()
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="✏️ Изменить название", callback_data="manage_rename"))
        builder.row(InlineKeyboardButton(text="❌ Удалить поставщика", callback_data="manage_delete"))
        builder.row(InlineKeyboardButton(text="🧹 Очистить баланс", callback_data="manage_clear_balance"))
        builder.row(InlineKeyboardButton(text="📅 Оплаты по периодам", callback_data="manage_pnl"))
        builder.row(InlineKeyboardButton(text="🔥 Полная очистка БД", callback_data="confirm_clear"))
        await message.answer("✅ Доступ разрешен.\n⚙️ **Меню управления данными:**", reply_markup=builder.as_markup(), parse_mode="Markdown")
    else:
        await message.answer("❌ Неверный ПИН-код. Попробуйте еще раз или нажмите /start для отмены.")

# --- P&L И ОТЧЕТЫ ПО ПЕРИОДАМ ---
@router.callback_query(F.data == "manage_pnl")
async def manage_pnl_list(callback: types.CallbackQuery):
    suppliers = await fb_get("suppliers")
    builder = InlineKeyboardBuilder()
    for s_id, s_data in suppliers.items():
        builder.add(InlineKeyboardButton(text=s_data['name'], callback_data=f"pnl_sup_{s_id}"))
    builder.adjust(2)
    await callback.message.edit_text("Выберите поставщика для детального отчета по оплатам (P&L):", reply_markup=builder.as_markup())

@router.callback_query(F.data.startswith("pnl_sup_"))
async def process_pnl_supplier(callback: types.CallbackQuery):
    sup_id = callback.data.split("_")[-1]
    suppliers = await fb_get("suppliers")
    debts = await fb_get("debts")
    
    s_name = suppliers.get(sup_id, {}).get('name', 'Неизвестен')
    
    # Фильтруем только оплаты по этому поставщику
    payments = [d for d in debts.values() if isinstance(d, dict) and d.get('supplier_id') == sup_id and d.get('is_paid') == 1]
    
    if not payments:
        await callback.message.edit_text(f"У поставщика **{s_name}** пока нет зафиксированных оплат.", parse_mode="Markdown")
        return

    # Группировка по месяцам для P&L
    pnl_report = {}
    for p in payments:
        date_str = p.get('created_at', '2024-01-01 00:00:00')
        month = date_str[:7] # YYYY-MM
        amount = float(p.get('amount', 0))
        pnl_report[month] = pnl_report.get(month, 0) + amount

    text = f"📊 **Отчет по оплатам (P&L): {s_name}**\n\n"
    total_all = 0
    for month, total in sorted(pnl_report.items(), reverse=True):
        text += f"📅 {month}: **{total:,.0f}** сум\n"
        total_all += total
    
    text += f"\n💰 Итого выплачено: **{total_all:,.0f}** сум"
    
    # Кнопка для получения детального Excel по этому поставщику
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="📥 Скачать детализацию (Excel)", callback_data=f"pnl_excel_{sup_id}"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data.startswith("pnl_excel_"))
async def process_pnl_excel(callback: types.CallbackQuery):
    sup_id = callback.data.split("_")[-1]
    suppliers = await fb_get("suppliers")
    debts = await fb_get("debts")
    s_name = suppliers.get(sup_id, {}).get('name', 'Supplier')
    
    payments = []
    for d in debts.values():
        if isinstance(d, dict) and d.get('supplier_id') == sup_id and d.get('is_paid') == 1:
            payments.append({
                "Дата": d.get('created_at', '-'),
                "Сумма (UZS)": d.get('amount', 0),
                "Тип": "ОПЛАТА"
            })
    
    if not payments:
        await callback.answer("Нет данных для выгрузки.")
        return

    df = pd.DataFrame(payments)
    filename = f"PNL_{s_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    df.to_excel(filename, index=False)
    
    file = FSInputFile(filename)
    await callback.message.answer_document(file, caption=f"📑 Детальная история оплат: {s_name}")
    os.remove(filename)
    await callback.answer()

# --- УПРАВЛЕНИЕ ПОСТАВЩИКАМИ ---
@router.callback_query(F.data == "manage_rename")
async def manage_rename_list(callback: types.CallbackQuery):
    suppliers = await fb_get("suppliers")
    builder = InlineKeyboardBuilder()
    for s_id, s_data in suppliers.items():
        builder.add(InlineKeyboardButton(text=s_data['name'], callback_data=f"rename_sup_{s_id}"))
    builder.adjust(2)
    await callback.message.edit_text("Выберите поставщика для переименования:", reply_markup=builder.as_markup())

@router.callback_query(F.data.startswith("rename_sup_"))
async def process_rename_select(callback: types.CallbackQuery, state: FSMContext):
    sup_id = callback.data.split("_")[-1]
    await state.update_data(sup_id=sup_id)
    await callback.message.edit_text("Введите новое название для поставщика:")
    await state.set_state(Form.edit_supplier_name)

@router.message(Form.edit_supplier_name)
async def process_rename_final(message: types.Message, state: FSMContext):
    data = await state.get_data()
    new_name = message.text.strip()
    await fb_put(f"suppliers/{data['sup_id']}/name", new_name)
    await message.answer(f"✅ Поставщик успешно переименован в '{new_name}'")
    await state.clear()

@router.callback_query(F.data == "manage_delete")
async def manage_delete_list(callback: types.CallbackQuery):
    suppliers = await fb_get("suppliers")
    builder = InlineKeyboardBuilder()
    for s_id, s_data in suppliers.items():
        builder.add(InlineKeyboardButton(text=s_data['name'], callback_data=f"delete_sup_{s_id}"))
    builder.adjust(2)
    await callback.message.edit_text("❌ ВНИМАНИЕ: При удалении поставщика удалятся ВСЕ его долги и оплаты.\nВыберите кого удалить:", reply_markup=builder.as_markup())

@router.callback_query(F.data.startswith("delete_sup_"))
async def process_delete_final(callback: types.CallbackQuery):
    sup_id = callback.data.split("_")[-1]
    debts = await fb_get("debts")
    if debts:
        for d_id, d_data in debts.items():
            if d_data['supplier_id'] == sup_id:
                await fb_delete(f"debts/{d_id}")
    await fb_delete(f"suppliers/{sup_id}")
    await callback.message.edit_text("✅ Поставщик и все его данные удалены.")

@router.callback_query(F.data == "manage_clear_balance")
async def manage_clear_balance_list(callback: types.CallbackQuery):
    suppliers = await fb_get("suppliers")
    builder = InlineKeyboardBuilder()
    for s_id, s_data in suppliers.items():
        builder.add(InlineKeyboardButton(text=s_data['name'], callback_data=f"clear_bal_{s_id}"))
    builder.adjust(2)
    await callback.message.edit_text("Выберите поставщика для очистки баланса (все записи станут 'оплаченными'):", reply_markup=builder.as_markup())

@router.callback_query(F.data.startswith("clear_bal_"))
async def process_clear_bal_final(callback: types.CallbackQuery):
    sup_id = callback.data.split("_")[-1]
    debts = await fb_get("debts")
    if debts:
        for d_id, d_data in debts.items():
            if d_data['supplier_id'] == sup_id:
                await fb_put(f"debts/{d_id}/is_paid", 1)
    await callback.message.edit_text("✅ Баланс поставщика очищен (все долги закрыты).")

# --- ПОСТАВЩИКИ ---
async def view_suppliers_msg(message: types.Message):
    suppliers = await fb_get("suppliers")
    if not suppliers:
        await message.answer("Список поставщиков пуст.")
        return
    text = "📦 **Список поставщиков:**\n\n" + "\n".join([f"• {s['name']}" for s in suppliers.values()])
    await message.answer(text, parse_mode="Markdown")

@router.callback_query(F.data == "view_suppliers")
async def view_suppliers(callback: types.CallbackQuery):
    await view_suppliers_msg(callback.message)
    await callback.answer()

async def add_supplier_start_msg(message: types.Message, state: FSMContext):
    await message.answer("Введите название нового поставщика:")
    await state.set_state(Form.add_supplier_name)

@router.callback_query(F.data == "add_supplier")
async def add_supplier_start(callback: types.CallbackQuery, state: FSMContext):
    await add_supplier_start_msg(callback.message, state)
    await callback.answer()

@router.message(Form.add_supplier_name)
async def process_supplier_name(message: types.Message, state: FSMContext):
    name = message.text.strip()
    suppliers = await fb_get("suppliers")
    if any(s['name'] == name for s in suppliers.values()):
        await message.answer(f"❌ Поставщик с таким названием уже существует.")
    else:
        await fb_post("suppliers", {"name": name})
        await message.answer(f"✅ Поставщик '{name}' успешно добавлен!", reply_markup=main_reply_keyboard())
    await state.clear()

def date_keyboard():
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📅 Сегодня", callback_data="date_today"))
    builder.row(InlineKeyboardButton(text="📅 Вчера", callback_data="date_yesterday"))
    builder.row(InlineKeyboardButton(text="⌨️ Ввести дату вручную", callback_data="date_manual"))
    return builder.as_markup()

# --- ДОЛГИ И БАЛАНС ---
async def view_debts_msg(message: types.Message):
    suppliers = await fb_get("suppliers")
    debts = await fb_get("debts")
    
    if not suppliers:
        await message.answer("Список поставщиков пуст.")
        return

    text = "💸 **Баланс по поставщикам (UZS):**\n\n"
    has_active = False

    for s_id, s_data in suppliers.items():
        total_debt = sum(float(d.get('amount', 0)) for d in debts.values() if isinstance(d, dict) and d.get('supplier_id') == s_id and d.get('is_paid') == 0)
        total_paid = sum(float(d.get('amount', 0)) for d in debts.values() if isinstance(d, dict) and d.get('supplier_id') == s_id and d.get('is_paid') == 1)
        
        remaining = total_debt - total_paid
        if remaining > 0:
            has_active = True
            text += f"• **{s_data['name']}**\n"
            text += f"   └ Долг: {total_debt:,.0f}\n"
            text += f"   └ Оплачено: {total_paid:,.0f}\n"
            text += f"   └ **Остаток: {remaining:,.0f}**\n\n"

    if not has_active:
        await message.answer("Активных долгов нет.")
    else:
        await message.answer(text, parse_mode="Markdown")

@router.callback_query(F.data == "view_debts")
async def view_debts(callback: types.CallbackQuery):
    await view_debts_msg(callback.message)
    await callback.answer()

async def add_debt_start_msg(message: types.Message, state: FSMContext):
    suppliers = await fb_get("suppliers")
    if not suppliers:
        await message.answer("❌ Сначала добавьте хотя бы одного поставщика.")
        return
    builder = InlineKeyboardBuilder()
    for s_id, s_data in suppliers.items():
        builder.add(InlineKeyboardButton(text=s_data['name'], callback_data=f"sel_sup_{s_id}"))
    builder.adjust(2)
    await message.answer("Выберите поставщика:", reply_markup=builder.as_markup())
    await state.set_state(Form.add_debt_supplier)

@router.callback_query(F.data == "add_debt")
async def add_debt_start(callback: types.CallbackQuery, state: FSMContext):
    await add_debt_start_msg(callback.message, state)
    await callback.answer()

@router.callback_query(F.data.startswith("sel_sup_"))
async def process_debt_supplier(callback: types.CallbackQuery, state: FSMContext):
    supplier_id = callback.data.split("_")[-1]
    await state.update_data(supplier_id=supplier_id)
    await callback.message.edit_text("Введите сумму долга (UZS):")
    await state.set_state(Form.add_debt_amount)

@router.message(Form.add_debt_amount)
async def process_debt_amount(message: types.Message, state: FSMContext):
    try:
        clean_amount = message.text.replace(' ', '').replace('.', '').replace(',', '.')
        amount = float(clean_amount)
        await state.update_data(amount=amount)
        await message.answer(f"Сумма {amount:,.0f} сум принята.\nВыберите дату операции:", reply_markup=date_keyboard())
        await state.set_state(Form.add_debt_date)
    except ValueError:
        await message.answer("❌ Ошибка! Введите корректное число.")

@router.callback_query(F.data.startswith("date_"), Form.add_debt_date)
async def process_debt_date_kb(callback: types.CallbackQuery, state: FSMContext):
    if callback.data == "date_manual":
        await callback.message.edit_text("Введите дату в формате ДД.ММ.ГГГГ (например, 10.04.2024):")
        return
    
    date_val = datetime.now()
    if callback.data == "date_yesterday":
        date_val = date_val.replace(day=date_val.day - 1)
    
    formatted_date = date_val.strftime('%Y-%m-%d %H:%M:%S')
    data = await state.get_data()
    
    await fb_post("debts", {
        "supplier_id": data['supplier_id'],
        "amount": data['amount'],
        "is_paid": 0,
        "created_at": formatted_date
    })
    
    await callback.message.edit_text(f"✅ Долг на сумму {data['amount']:,.0f} сум записан за {formatted_date[:10]}!")
    await state.clear()

@router.message(Form.add_debt_date)
async def process_debt_date_manual(message: types.Message, state: FSMContext):
    try:
        date_obj = datetime.strptime(message.text.strip(), "%d.%m.%Y")
        formatted_date = date_obj.strftime('%Y-%m-%d %H:%M:%S')
        data = await state.get_data()
        
        await fb_post("debts", {
            "supplier_id": data['supplier_id'],
            "amount": data['amount'],
            "is_paid": 0,
            "created_at": formatted_date
        })
        
        await message.answer(f"✅ Долг записан за {message.text.strip()}!", reply_markup=main_reply_keyboard())
        await state.clear()
    except ValueError:
        await message.answer("❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ (например, 10.04.2024):")

# --- ОПЛАТА ---
async def add_payment_start_msg(message: types.Message, state: FSMContext):
    suppliers = await fb_get("suppliers")
    if not suppliers:
        await message.answer("❌ Сначала добавьте хотя бы одного поставщика.")
        return
    builder = InlineKeyboardBuilder()
    for s_id, s_data in suppliers.items():
        builder.add(InlineKeyboardButton(text=s_data['name'], callback_data=f"pay_sup_{s_id}"))
    builder.adjust(2)
    await message.answer("Выберите поставщика для оплаты:", reply_markup=builder.as_markup())
    await state.set_state(Form.add_payment_supplier)

@router.callback_query(F.data.startswith("pay_sup_"))
async def process_payment_supplier(callback: types.CallbackQuery, state: FSMContext):
    supplier_id = callback.data.split("_")[-1]
    await state.update_data(supplier_id=supplier_id)
    await callback.message.edit_text("Введите сумму оплаты (UZS):")
    await state.set_state(Form.add_payment_amount)

@router.message(Form.add_payment_amount)
async def process_payment_amount(message: types.Message, state: FSMContext):
    try:
        clean_amount = message.text.replace(' ', '').replace('.', '').replace(',', '.')
        amount = float(clean_amount)
        await state.update_data(amount=amount)
        await message.answer(f"Сумма оплаты {amount:,.0f} сум принята.\nВыберите дату операции:", reply_markup=date_keyboard())
        await state.set_state(Form.add_payment_date)
    except ValueError:
        await message.answer("❌ Ошибка! Введите корректное число.")

@router.callback_query(F.data.startswith("date_"), Form.add_payment_date)
async def process_payment_date_kb(callback: types.CallbackQuery, state: FSMContext):
    if callback.data == "date_manual":
        await callback.message.edit_text("Введите дату в формате ДД.ММ.ГГГГ (например, 10.04.2024):")
        return
    
    date_val = datetime.now()
    if callback.data == "date_yesterday":
        date_val = date_val.replace(day=date_val.day - 1)
    
    formatted_date = date_val.strftime('%Y-%m-%d %H:%M:%S')
    data = await state.get_data()
    
    await fb_post("debts", {
        "supplier_id": data['supplier_id'],
        "amount": data['amount'],
        "is_paid": 1,
        "created_at": formatted_date
    })
    
    await callback.message.edit_text(f"✅ Оплата {data['amount']:,.0f} сум записана за {formatted_date[:10]}!")
    await state.clear()

@router.message(Form.add_payment_date)
async def process_payment_date_manual(message: types.Message, state: FSMContext):
    try:
        date_obj = datetime.strptime(message.text.strip(), "%d.%m.%Y")
        formatted_date = date_obj.strftime('%Y-%m-%d %H:%M:%S')
        data = await state.get_data()
        
        await fb_post("debts", {
            "supplier_id": data['supplier_id'],
            "amount": data['amount'],
            "is_paid": 1,
            "created_at": formatted_date
        })
        
        await message.answer(f"✅ Оплата записана за {message.text.strip()}!", reply_markup=main_reply_keyboard())
        await state.clear()
    except ValueError:
        await message.answer("❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ (например, 10.04.2024):")

# --- ЭКСПОРТ ---
async def export_to_excel_msg(message: types.Message):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    print("DEBUG: Начинаем экспорт в Excel...")
    try:
        suppliers = await fb_get("suppliers")
        debts = await fb_get("debts")
        
        print(f"DEBUG: Поставщиков: {len(suppliers)}, Долгов: {len(debts)}")
        
        if not suppliers:
            await message.answer("База данных пуста.")
            return

        # Сводные данные
        summary_data = []
        for s_id, s_data in suppliers.items():
            total_debt = sum(float(d.get('amount', 0)) for d in debts.values() if isinstance(d, dict) and d.get('supplier_id') == s_id and d.get('is_paid') == 0)
            total_paid = sum(float(d.get('amount', 0)) for d in debts.values() if isinstance(d, dict) and d.get('supplier_id') == s_id and d.get('is_paid') == 1)
            summary_data.append({
                "Поставщик": s_data['name'],
                "Общий долг": total_debt,
                "Всего оплачено": total_paid,
                "Остаток к оплате": total_debt - total_paid
            })
        df_summary = pd.DataFrame(summary_data)

        # История операций
        history_data = []
        if debts:
            for d_data in debts.values():
                if not isinstance(d_data, dict) or 'supplier_id' not in d_data:
                    continue
                s_name = suppliers.get(d_data['supplier_id'], {}).get('name', 'Удален')
                history_data.append({
                    "Поставщик": s_name,
                    "Сумма (UZS)": d_data.get('amount', 0),
                    "Тип": "ОПЛАТА" if d_data.get('is_paid') == 1 else "ДОЛГ",
                    "Дата": d_data.get('created_at', '-')
                })
        df_history = pd.DataFrame(history_data) if history_data else pd.DataFrame(columns=["Поставщик", "Сумма (UZS)", "Тип", "Дата"])

        filename = f"Debt_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"DEBUG: Создаем файл {filename}")
        
        # Сортировка истории по дате
        if not df_history.empty:
            df_history['Дата'] = pd.to_datetime(df_history['Дата'], errors='coerce')
            df_history = df_history.sort_values(by='Дата', ascending=False)
            df_history['Дата'] = df_history['Дата'].dt.strftime('%d.%m.%Y %H:%M')

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Сводный баланс', index=False)
            df_history.to_excel(writer, sheet_name='История операций', index=False)
            
            workbook = writer.book
            
            # Стили для всех листов
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            align_center = Alignment(horizontal='center')

            for sheet_name in ['Сводный баланс', 'История операций']:
                ws = writer.sheets[sheet_name]
                # Оформление шапки
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = border
                
                # Оформление данных
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        # Форматирование сумм
                        if sheet_name == 'Сводный баланс' and cell.column in [2, 3, 4]:
                            cell.number_format = '#,##0 "сум"'
                        if sheet_name == 'История операций' and cell.column == 2:
                            cell.number_format = '#,##0 "сум"'
                
                # Автоподбор ширины
                for i, col in enumerate(ws.columns, 1):
                    ws.column_dimensions[col[0].column_letter].width = 25

        print(f"DEBUG: Файл создан успешно. Отправляем в Telegram...")
        file = FSInputFile(filename)
        await message.answer_document(file, caption=f"📊 Балансовый отчет\n💰 Остаток: {df_summary['Остаток к оплате'].sum():,.0f} сум")
        
        if os.path.exists(filename):
            os.remove(filename)
            print(f"DEBUG: Файл {filename} удален после отправки.")
            
    except Exception as e:
        print(f"ERROR в export_to_excel_msg: {e}")
        await message.answer(f"❌ Произошла ошибка при экспорте: {e}")

# --- ОЧИСТКА ---
async def confirm_clear_msg(message: types.Message):
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="✅ Да, очистить всё", callback_data="clear_database"))
    builder.add(InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_clear"))
    await message.answer("❗ ВНИМАНИЕ: Вы уверены, что хотите безвозвратно удалить все данные из базы?", 
                         reply_markup=builder.as_markup())

@router.callback_query(F.data == "confirm_clear")
async def confirm_clear(callback: types.CallbackQuery):
    await confirm_clear_msg(callback.message)
    await callback.answer()

@router.callback_query(F.data == "clear_database")
async def clear_database(callback: types.CallbackQuery):
    await fb_delete("suppliers")
    await fb_delete("debts")
    await callback.message.edit_text("🔥 База данных полностью очищена.")

@router.callback_query(F.data == "cancel_clear")
async def cancel_clear(callback: types.CallbackQuery):
    await callback.message.edit_text("Очистка отменена.")

# Запуск
async def main():
    dp = Dispatcher()
    dp.include_router(router)
    print("Бот запущен...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        print("Бот остановлен.")
